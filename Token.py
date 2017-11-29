# coding=utf-8
from openpyxl import load_workbook
import re,collections

#打开一个workbook
wb = load_workbook(filename="/home/monkeys/下载/words.xlsx")

#获取当前活跃的worksheet,默认就是第一个worksheet
#ws = wb.active

#当然也可以使用下面的方法

#获取所有表格(worksheet)的名字
sheets = wb.get_sheet_names()
#第一个表格的名称
sheet_first = sheets[0]
#获取特定的worksheet
ws = wb.get_sheet_by_name(sheet_first)

#获取表格所有行和列，两者都是可迭代的
rows = ws.rows
columns = ws.columns

#储存所有单词的列表
words_list = []

#迭代所有的行
for row in rows:
 line = [col.value for col in row]
 for w in line[12:]:
  try:
   words_list.append(w.strip())
  except:
   pass
words_frequency = collections.Counter(words_list[1:])
# print words_frequency
data = sorted(words_frequency.iteritems(),key= lambda s:s[1],reverse=True)
for i in data:
 print i[0],i[1]
#通过坐标读取值
# print ws.cell('A1').value # A表示列,1表示行 u'\u9632\u706b\u5899'
# print ws.cell(row=1, column=1).value